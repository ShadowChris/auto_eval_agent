from __future__ import annotations

import json
from io import BytesIO

import pandas as pd
import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from auto_eval.web import server
from auto_eval.web.operation_comparison_import import (
    import_operation_comparison_file,
)
from auto_eval.web.server import (
    OperationComparisonAnalyzeReq,
    OperationComparisonSourceReq,
)


def test_import_comparison_jsonl_reads_nested_evaluation() -> None:
    content = "\n".join([
        json.dumps({
            "id": "case_1",
            "index": "simple_001",
            "case_id": "c1",
            "query": "打开蓝牙",
            "context": "测试环境",
            "evaluation": {
                "correctness": "ok",
                "issue_types": [],
                "rationale": "已经打开",
            },
        }, ensure_ascii=False),
        json.dumps({
            "id": "case_2",
            "index": "simple_002",
            "case_id": "c2",
            "query": "关闭定位",
            "evaluation": {
                "correctness": "nok",
                "issue_types": ["应执行目标未执行"],
            },
        }, ensure_ascii=False),
    ]).encode("utf-8")

    source = import_operation_comparison_file("结果.jsonl", content)

    assert source["summary"]["raw_count"] == 2
    assert source["summary"]["valid_count"] == 2
    assert source["task_id"] == source["source_id"]
    assert source["group_name"] == "结果.jsonl"
    assert source["mapping"]["index"] == "index"
    assert source["rows"][0]["index"] == "simple_001"
    assert source["mapping"]["correctness"] == "evaluation.correctness"
    assert source["rows"][1]["result"]["issue_types"] == ["应执行目标未执行"]
    assert source["rows"][0]["export"]["context"] == "测试环境"


def test_import_comparison_excel_prefers_result_sheet() -> None:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame([{"query": "错误工作表"}]).to_excel(
            writer,
            sheet_name="数据集明细",
            index=False,
        )
        pd.DataFrame([{
            "item_id": "i1",
            "index": "simple_001",
            "case_id": "c1",
            "query": "打开蓝牙",
            "correctness": "ok",
            "issue_types": "",
        }]).to_excel(writer, sheet_name="逐题结果", index=False)

    source = import_operation_comparison_file("结果.xlsx", output.getvalue())

    assert source["summary"]["sheet"] == "逐题结果"
    assert source["summary"]["valid_count"] == 1
    assert source["rows"][0]["query"] == "打开蓝牙"


def test_import_comparison_csv_maps_chinese_columns_and_warnings() -> None:
    content = (
        "index,case_id,操作意图,完成判定,问题类型\n"
        "simple_001,c1,打开蓝牙,OK,路径冗余；重复操作\n"
        "simple_002,c2,关闭定位,unknown,\n"
    ).encode("utf-8-sig")

    with TestClient(server.app) as client:
        response = client.post(
            "/api/operation/comparison/import",
            files={"file": ("实验组.csv", content, "text/csv")},
        )

    assert response.status_code == 200
    source = response.json()
    assert source["mapping"]["query"] == "操作意图"
    assert source["summary"]["valid_count"] == 1
    assert source["summary"]["invalid_count"] == 1
    assert source["summary"]["warning_count"] == 1
    assert source["rows"][0]["result"]["issue_types"] == ["路径冗余", "重复操作"]


def test_import_comparison_rejects_file_without_valid_correctness() -> None:
    content = "case_id,query,correctness\nc1,打开蓝牙,partial\n".encode("utf-8")

    with pytest.raises(ValueError, match="未识别到有效 correctness"):
        import_operation_comparison_file("旧标准.csv", content)


def _history_snapshot() -> dict:
    return {
        "task_id": "history_1",
        "dataset_name": "历史对照.jsonl",
        "mode": "operation",
        "status": "done",
        "options": {},
        "items": [{
            "id": "i1",
            "case_id": "c1",
            "query": "打开蓝牙",
            "source_data": {"index": "simple_001"},
        }],
        "results": [{"index": 0, "correctness": "nok", "issue_types": ["应执行目标未执行"]}],
        "summary": {},
    }


def test_comparison_analysis_mixes_history_and_upload(monkeypatch) -> None:
    imported = import_operation_comparison_file(
        "实验.jsonl",
        json.dumps({
            "id": "i1",
            "index": "simple_001",
            "case_id": "c1",
            "query": "打开蓝牙",
            "evaluation": {"correctness": "ok", "issue_types": []},
        }, ensure_ascii=False).encode("utf-8"),
    )
    monkeypatch.setattr(server, "get_live_task", lambda _task_id: None)
    monkeypatch.setattr(
        server,
        "load_snapshot",
        lambda task_id: _history_snapshot() if task_id == "history_1" else None,
    )
    request = OperationComparisonAnalyzeReq(
        sources=[
            OperationComparisonSourceReq(
                source_id="history_1",
                source_type="history",
                task_id="history_1",
                group_name="历史对照",
            ),
            OperationComparisonSourceReq(**imported),
        ],
        control_source_id="history_1",
    )

    payload = server.api_operation_comparison_analyze(request)

    assert payload["group_count"] == 2
    assert payload["all_groups_common_valid_count"] == 1
    assert payload["groups"][0]["group_label"] == "对照组"
    assert payload["groups"][1]["group_label"] == "实验组A"
    assert payload["pairwise"][0]["to_ok_count"] == 1


def test_comparison_export_mixes_history_and_upload(monkeypatch) -> None:
    imported = import_operation_comparison_file(
        "实验.jsonl",
        json.dumps({
            "id": "i1",
            "index": "simple_001",
            "case_id": "c1",
            "query": "打开蓝牙",
            "sessionid": "session-a",
            "evaluation": {"correctness": "ok", "issue_types": []},
        }, ensure_ascii=False).encode("utf-8"),
    )
    monkeypatch.setattr(server, "get_live_task", lambda _task_id: None)
    monkeypatch.setattr(
        server,
        "load_snapshot",
        lambda task_id: _history_snapshot() if task_id == "history_1" else None,
    )
    request = OperationComparisonAnalyzeReq(
        sources=[
            OperationComparisonSourceReq(
                source_id="history_1",
                source_type="history",
                task_id="history_1",
                group_name="历史对照",
            ),
            OperationComparisonSourceReq(**imported),
        ],
        control_source_id="history_1",
    )

    response = server.api_operation_comparison_export(request)
    workbook = load_workbook(BytesIO(response.body), read_only=True)

    assert workbook.sheetnames == ["对比概览", "Issue Type对比", "逐题横向对比"]
    detail = workbook["逐题横向对比"]
    header_rows = detail.iter_rows()
    group_headers = [cell.value for cell in next(header_rows)]
    field_headers = [cell.value for cell in next(header_rows)]
    assert "实验组A：实验" in group_headers
    assert field_headers.count("sessionid") == 2


def test_comparison_module_ui_is_parallel_workspace() -> None:
    html = (server.STATIC_DIR / "index.html").read_text(encoding="utf-8")
    js = (server.STATIC_DIR / "app.js").read_text(encoding="utf-8")

    assert "任务类对比分析" in html
    assert "taskModule==='comparison'" in html
    assert "上传评估结果集" in html
    assert "加入对比分析" in html
    assert 'fetch("/api/operation/comparison/import"' in js
    assert 'fetch("/api/operation/comparison/analyze"' in js
    assert 'fetch("/api/operation/comparison/export"' in js
