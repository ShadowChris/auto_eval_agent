"""按文件名中的设备编号分组合并 CSV/XLSX。

输入目录中的 CSV 或 XLSX 文件名应以设备编号结尾，例如：
``版本11.6.3_20260817_DEVICE_EXP_001.xlsx``。同一设备可以有多个
输入文件，脚本会把它们全部归入该设备所在的组，在每行追加
``device_id``，并按“序号”列自然升序排列。XLSX 默认读取第一个工作表。

示例：

    python scripts/csv_merge_by_device_group.py data/0817/0817-CSV \
      --group "实验组:DEVICE_EXP_001;DEVICE_EXP_002" \
      --group "对照组:DEVICE_CTRL_001;DEVICE_CTRL_002" \
      --output-name "0817设置众测"

也可以把多行分组定义写入文本文件，再使用 ``--groups-file`` 传入。
"""

from __future__ import annotations

import argparse
import csv
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Sequence


DEFAULT_SEQUENCE_COLUMN = "序号"
DEFAULT_DEVICE_COLUMN = "device_id"
DEFAULT_ENCODING = "utf-8-sig"
SUPPORTED_INPUT_SUFFIXES = {".csv", ".xlsx"}


@dataclass(frozen=True)
class GroupMergeResult:
    group_name: str
    device_ids: tuple[str, ...]
    source_files: tuple[Path, ...]
    row_count: int
    output_path: Path


@dataclass(frozen=True)
class MergeResult:
    input_dir: Path
    output_dir: Path
    groups: tuple[GroupMergeResult, ...]
    unmatched_files: tuple[Path, ...]


def parse_group_definition(value: str) -> tuple[str, tuple[str, ...]]:
    """解析“组名:设备1;设备2”格式，兼容中文冒号和分号。"""
    normalized = value.strip()
    match = re.match(r"^([^:：]+)[:：](.+)$", normalized)
    if not match:
        raise ValueError(f"分组格式错误，应为‘组名:设备1;设备2’：{value}")
    group_name = match.group(1).strip()
    device_ids = tuple(
        device_id.strip()
        for device_id in re.split(r"[;；]", match.group(2))
        if device_id.strip()
    )
    if not group_name:
        raise ValueError("组名不能为空")
    if not device_ids:
        raise ValueError(f"分组“{group_name}”至少需要一个设备编号")
    if len(set(device_ids)) != len(device_ids):
        raise ValueError(f"分组“{group_name}”包含重复设备编号")
    return group_name, device_ids


def parse_group_definitions(values: Iterable[str]) -> dict[str, tuple[str, ...]]:
    groups: dict[str, tuple[str, ...]] = {}
    device_owner: dict[str, str] = {}
    for value in values:
        text = value.strip()
        if not text or text.startswith("#"):
            continue
        group_name, device_ids = parse_group_definition(text)
        if group_name in groups:
            raise ValueError(f"组名重复：{group_name}")
        for device_id in device_ids:
            previous = device_owner.get(device_id)
            if previous is not None:
                raise ValueError(
                    f"设备编号 {device_id} 同时属于“{previous}”和“{group_name}”"
                )
            device_owner[device_id] = group_name
        groups[group_name] = device_ids
    if not groups:
        raise ValueError("至少需要一个有效分组")
    return groups


def _safe_filename_component(value: str, field: str) -> str:
    text = value.strip()
    if not text:
        raise ValueError(f"{field}不能为空")
    if text in {".", ".."} or any(char in text for char in "/\\\0"):
        raise ValueError(f"{field}不能包含路径分隔符：{value}")
    return text.removesuffix(".csv")


def _filename_device_id(path: Path, device_ids: Sequence[str]) -> str | None:
    matches = [
        device_id
        for device_id in device_ids
        if re.search(rf"(?:^|[_-]){re.escape(device_id)}$", path.stem)
    ]
    if len(matches) > 1:
        raise ValueError(f"文件名同时匹配多个设备编号：{path.name}")
    return matches[0] if matches else None


def _natural_key(value: object) -> tuple[tuple[int, object], ...]:
    parts = re.split(r"(\d+)", str(value or "").strip().casefold())
    return tuple(
        (0, int(part)) if part.isdigit() else (1, part)
        for part in parts
        if part != ""
    )


def _validate_fieldnames(path: Path, values: Sequence[object]) -> list[str]:
    fieldnames = [str(value).strip() if value is not None else "" for value in values]
    while fieldnames and not fieldnames[-1]:
        fieldnames.pop()
    if not fieldnames:
        raise ValueError(f"表格缺少表头：{path}")
    if not all(fieldnames):
        raise ValueError(f"表格存在空列名：{path}")
    if len(set(fieldnames)) != len(fieldnames):
        raise ValueError(f"表格存在重复列名：{path}")
    return fieldnames


def _read_csv(path: Path, encoding: str) -> tuple[list[str], list[dict[str, object]]]:
    try:
        with path.open("r", encoding=encoding, newline="") as source:
            reader = csv.DictReader(source)
            if reader.fieldnames is None:
                raise ValueError(f"CSV 缺少表头：{path}")
            source_fieldnames = list(reader.fieldnames)
            fieldnames = _validate_fieldnames(path, source_fieldnames)
            if len(fieldnames) != len(source_fieldnames):
                raise ValueError(f"CSV 存在空列名：{path}")
            rows: list[dict[str, object]] = []
            for line_number, row in enumerate(reader, start=2):
                if None in row:
                    raise ValueError(
                        f"CSV 第 {line_number} 行字段数超过表头：{path}"
                    )
                rows.append({
                    fieldnames[index]: row.get(source_name, "")
                    for index, source_name in enumerate(source_fieldnames)
                })
    except UnicodeDecodeError as exc:
        raise ValueError(f"无法按 {encoding} 解码 CSV：{path}") from exc
    return fieldnames, rows


def _read_xlsx(path: Path) -> tuple[list[str], list[dict[str, object]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError(
            "读取 XLSX 需要数据处理依赖，请运行 "
            'python -m pip install -e ".[data]"'
        ) from exc

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"无法读取 XLSX：{path}（{exc}）") from exc
    try:
        if not workbook.worksheets:
            raise ValueError(f"XLSX 不包含工作表：{path}")
        worksheet = workbook.worksheets[0]
        values = worksheet.iter_rows(values_only=True)
        header = next(values, None)
        if header is None:
            raise ValueError(f"XLSX 第一个工作表为空：{path}")
        fieldnames = _validate_fieldnames(path, header)
        rows: list[dict[str, object]] = []
        for row_number, row in enumerate(values, start=2):
            extra_values = row[len(fieldnames):]
            if any(value not in (None, "") for value in extra_values):
                raise ValueError(
                    f"XLSX 第 {row_number} 行字段数超过表头：{path}"
                )
            normalized = list(row[:len(fieldnames)])
            normalized.extend([""] * (len(fieldnames) - len(normalized)))
            if all(value in (None, "") for value in normalized):
                continue
            rows.append({
                fieldnames[index]: value if value is not None else ""
                for index, value in enumerate(normalized)
            })
        return fieldnames, rows
    finally:
        workbook.close()


def _read_table(
    path: Path,
    encoding: str,
) -> tuple[list[str], list[dict[str, object]]]:
    if path.suffix.lower() == ".csv":
        return _read_csv(path, encoding)
    if path.suffix.lower() == ".xlsx":
        return _read_xlsx(path)
    raise ValueError(f"不支持的输入格式：{path}")


def merge_csv_groups(
    input_dir: Path,
    groups: dict[str, Sequence[str]],
    *,
    output_name: str,
    output_dir: Path | None = None,
    sequence_column: str = DEFAULT_SEQUENCE_COLUMN,
    device_column: str = DEFAULT_DEVICE_COLUMN,
    encoding: str = DEFAULT_ENCODING,
) -> MergeResult:
    input_dir = input_dir.expanduser().resolve()
    if not input_dir.is_dir():
        raise FileNotFoundError(f"输入目录不存在：{input_dir}")
    normalized_output_name = _safe_filename_component(output_name, "输出名称")
    normalized_groups = parse_group_definitions(
        f"{group_name}:{';'.join(device_ids)}"
        for group_name, device_ids in groups.items()
    )
    output_dir = (
        output_dir.expanduser().resolve()
        if output_dir is not None
        else input_dir / "merged_output"
    )
    output_dir.mkdir(parents=True, exist_ok=True)

    input_files = sorted(
        (
            path
            for path in input_dir.iterdir()
            if path.is_file()
            and path.suffix.lower() in SUPPORTED_INPUT_SUFFIXES
            and not path.name.startswith("~$")
        ),
        key=lambda path: path.name.casefold(),
    )
    if not input_files:
        raise ValueError(f"输入目录没有 CSV 或 XLSX 文件：{input_dir}")

    all_device_ids = tuple(
        device_id
        for device_ids in normalized_groups.values()
        for device_id in device_ids
    )
    files_by_device: dict[str, list[Path]] = {device_id: [] for device_id in all_device_ids}
    unmatched_files: list[Path] = []
    for path in input_files:
        device_id = _filename_device_id(path, all_device_ids)
        if device_id is None:
            unmatched_files.append(path)
        else:
            files_by_device[device_id].append(path)

    missing_devices = [
        device_id for device_id, paths in files_by_device.items() if not paths
    ]
    if missing_devices:
        raise ValueError(
            "以下设备编号没有匹配到以该编号结尾的 CSV/XLSX："
            + "、".join(missing_devices)
        )

    results: list[GroupMergeResult] = []
    for group_name, device_ids in normalized_groups.items():
        safe_group_name = _safe_filename_component(group_name, "组名")
        source_files = tuple(
            path
            for device_id in device_ids
            for path in files_by_device[device_id]
        )
        fieldnames: list[str] = []
        merged_rows: list[dict[str, object]] = []
        for device_id in device_ids:
            for path in files_by_device[device_id]:
                source_fields, rows = _read_table(path, encoding)
                for field in source_fields:
                    if field != device_column and field not in fieldnames:
                        fieldnames.append(field)
                for row in rows:
                    row[device_column] = device_id
                    merged_rows.append(row)

        if sequence_column not in fieldnames:
            raise ValueError(
                f"分组“{group_name}”的 CSV 缺少排序列：{sequence_column}"
            )
        fieldnames = [field for field in fieldnames if field != device_column]
        fieldnames.append(device_column)
        merged_rows.sort(
            key=lambda row: (
                not bool(str(row.get(sequence_column) or "").strip()),
                _natural_key(row.get(sequence_column)),
            )
        )

        output_path = output_dir / f"{normalized_output_name}_{safe_group_name}.csv"
        if output_path in source_files:
            raise ValueError(f"输出文件不能覆盖输入文件：{output_path}")
        temporary_path = output_path.with_suffix(".csv.tmp")
        try:
            with temporary_path.open("w", encoding=encoding, newline="") as output:
                writer = csv.DictWriter(
                    output,
                    fieldnames=fieldnames,
                    extrasaction="ignore",
                    lineterminator="\n",
                )
                writer.writeheader()
                writer.writerows(merged_rows)
            temporary_path.replace(output_path)
        finally:
            temporary_path.unlink(missing_ok=True)

        results.append(GroupMergeResult(
            group_name=group_name,
            device_ids=tuple(device_ids),
            source_files=source_files,
            row_count=len(merged_rows),
            output_path=output_path,
        ))

    return MergeResult(
        input_dir=input_dir,
        output_dir=output_dir,
        groups=tuple(results),
        unmatched_files=tuple(unmatched_files),
    )


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="按设备编号分组合并 CSV/XLSX")
    parser.add_argument(
        "input_dir",
        type=Path,
        help="待合并 CSV/XLSX 所在目录（不递归；XLSX 读取第一个工作表）",
    )
    parser.add_argument(
        "--group",
        action="append",
        default=[],
        help="分组定义：组名:设备1;设备2。可重复传入",
    )
    parser.add_argument(
        "--groups-file",
        type=Path,
        help="UTF-8 分组定义文本，一行一个分组；空行和 # 注释会忽略",
    )
    parser.add_argument(
        "--output-name",
        required=True,
        help="输出文件基础名称；实际文件名为 <名称>_<组名>.csv",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        help="输出目录；默认为输入目录下的 merged_output 文件夹",
    )
    parser.add_argument(
        "--sequence-column",
        default=DEFAULT_SEQUENCE_COLUMN,
        help="排序列名，默认‘序号’",
    )
    parser.add_argument(
        "--device-column",
        default=DEFAULT_DEVICE_COLUMN,
        help="新增设备列名，默认 device_id",
    )
    parser.add_argument(
        "--encoding",
        default=DEFAULT_ENCODING,
        help="CSV 输入和合并结果编码，默认 utf-8-sig；不影响 XLSX",
    )
    return parser


def main(argv: list[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    definitions = list(args.group)
    if args.groups_file is not None:
        groups_file = args.groups_file.expanduser().resolve()
        if not groups_file.is_file():
            parser.error(f"分组文件不存在：{groups_file}")
        definitions.extend(groups_file.read_text(encoding="utf-8-sig").splitlines())
    try:
        groups = parse_group_definitions(definitions)
        result = merge_csv_groups(
            args.input_dir,
            groups,
            output_name=args.output_name,
            output_dir=args.output_dir,
            sequence_column=args.sequence_column,
            device_column=args.device_column,
            encoding=args.encoding,
        )
    except (FileNotFoundError, OSError, ValueError) as exc:
        parser.error(str(exc))

    for group in result.groups:
        print(
            f"[{group.group_name}] 设备 {len(group.device_ids)} 个，"
            f"源文件 {len(group.source_files)} 个，数据 {group.row_count} 行"
        )
        print(f"  输出：{group.output_path}")
    if result.unmatched_files:
        print(f"未参与合并的 CSV：{len(result.unmatched_files)} 个")
        for path in result.unmatched_files:
            print(f"  - {path.name}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
